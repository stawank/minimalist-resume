from fastapi import FastAPI, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from langchain_community.embeddings.fastembed import FastEmbedEmbeddings
from langchain_community.vectorstores import Chroma
from langchain_anthropic import ChatAnthropic
from pydantic import BaseModel
from datetime import datetime
from dotenv import load_dotenv
from functools import lru_cache
import openpyxl
import json
import os
import psutil

load_dotenv()
def get_system_stats():
    try:
        # Get CPU Temp (Specific to Raspberry Pi / Linux)
        with open("/sys/class/thermal/thermal_zone0/temp", "r") as f:
            temp = float(f.read()) / 1000.0
        
        cpu_usage = psutil.cpu_percent()
        ram_usage = psutil.virtual_memory().percent
        return f" (System Status: CPU {temp:.1f}°C, CPU Usage {cpu_usage}%, RAM {ram_usage}%)"
    except:
        return "" # Fallback if not running on Linux/Pi

DB_PATH        = os.getenv("DB_PATH")
LOG_PATH       = os.getenv("LOG_PATH")
VISITOR_LOG    = os.getenv("VISITOR_LOG", "/home/stawan/minimalist-resume/backend/visitors.xlsx")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
BLOCKED_SOURCES = {
    "Arbeitszeugnis_FPraktikum_Kulkarni.pdf",
    "Arbeitszeugnis_PPraktikum_Kulkarni.pdf",
    "Arbeitszeugnis_Werkstudent_Kulkarni.pdf",
    "Zwischenzeugnis_EDAG.pdf",
    "InterimReport_EDAG.pdf",
    "GECA_TRANSCRIPT.pdf",
    "FINAL DEGREE CERTIFICATE (1).pdf",
}

print("Loading resume memory...")
embeddings = FastEmbedEmbeddings(model_name="BAAI/bge-small-en-v1.5")
vector_db  = Chroma(persist_directory=DB_PATH, embedding_function=embeddings)
print(f"Loaded {vector_db._collection.count()} chunks")

print("Loading Anthropic...")
llm = ChatAnthropic(
    model="claude-haiku-4-5",
    anthropic_api_key=ANTHROPIC_API_KEY,
    temperature=0.0,
)
print("LLM Model ready!")

SYSTEM_PROMPT = """You are a professional AI assistant for Stawan Chandrashekhar Kulkarni.
You are currently running on a self-hosted Raspberry Pi 4 in Regensburg, Germany.

STRICT RULES:
- Answer ONLY using facts from the context. If not found, say "I don't have that information."
- Respond in the SAME language as the question (English or German).
- Be concise (max 3 sentences).
- If asked about the website or how you are running, explain you are a RAG-based LLM (Claude) self-hosted on Stawan's Raspberry Pi using FastAPI and Tailscale.
- Bio: Stawan is an Automotive Engineer with 2+ years of exp, focused on ADAS, sensor fusion (C++, Python, CAPL). He has a B1 level in German."""

def extract_text(response):
    content = response.content
    if isinstance(content, list):
        for item in content:
            if isinstance(item, dict) and item.get("type") == "text":
                return item["text"].strip()
            if isinstance(item, str):
                return item.strip()
        return ""
    return content.strip()

def build_prompt(context, question, history=[]):
    history_text = "\n".join(history[-4:]) + "\n" if history else ""
    stats = get_system_stats()
    #We append the stats to the system prompt so the AI knows the current "state" of its brain
    dynamic_system_prompt = SYSTEM_PROMPT + f"\n\nCURRENT_HARDWARE_STATS: {stats}"
    
    return f"{dynamic_system_prompt}\n\nContext:\n{context}\n\n{history_text}Question: {question}\nAnswer:"

def log_to_excel(path, headers, row):
    try:
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(headers)
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)
        ws.append(row)
        wb.save(path)
    except Exception as e:
        print(f"Excel log error: {e}")

def log_question(question, answer):
    log_to_excel(LOG_PATH, ["Timestamp", "Question", "Answer"],
        [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), question, answer])

def log_visitor(ip, user_agent):
    log_to_excel(VISITOR_LOG, ["Timestamp", "IP", "User Agent"],
        [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ip, user_agent])

def retrieve_context(question, history=[]):
    search_query = " ".join(history[-4:]) + " " + question if history else question
    docs = vector_db.similarity_search(search_query, k=5)
    safe_docs = [d for d in docs if d.metadata.get("source", "") not in BLOCKED_SOURCES]
    context = "\n\n".join([d.page_content for d in safe_docs])
    return context, safe_docs

@lru_cache(maxsize=50)
def cached_answer(question):
    context, safe_docs = retrieve_context(question)
    if not safe_docs:
        return "I don't have that information."
    prompt = build_prompt(context, question)
    return extract_text(llm.invoke(prompt))

class ChatRequest(BaseModel):
    question: str
    history: list[str] = []

class ChatResponse(BaseModel):
    answer: str

app = FastAPI(title="Resume Chatbot")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.post("/chat/stream")
def chat_stream(req: ChatRequest, request: Request):
    ip = request.headers.get("X-Real-IP", request.client.host)
    log_visitor(ip, request.headers.get("User-Agent", "unknown"))

    if not req.history:
        cached = cached_answer(req.question.lower().strip())
        def serve_cached():
            yield f"data: {json.dumps({'token': cached})}\n\n"
            yield "data: [DONE]\n\n"
            log_question(req.question, cached)
        return StreamingResponse(serve_cached(), media_type="text/event-stream")

    context, safe_docs = retrieve_context(req.question, req.history)

    if not safe_docs:
        def empty():
            yield f"data: {json.dumps({'token': 'I dont have that information.'})}\n\n"
            yield "data: [DONE]\n\n"
        return StreamingResponse(empty(), media_type="text/event-stream")

    prompt = build_prompt(context, req.question, req.history)

    def generate():
        full_answer = ""
        for chunk in llm.stream(prompt):
            raw = chunk.content
            if isinstance(raw, list):
                token = next((i["text"] for i in raw if isinstance(i, dict) and i.get("type") == "text"), "")
            else:
                token = raw or ""
            if token:
                full_answer += token
                yield f"data: {json.dumps({'token': token})}\n\n"
        log_question(req.question, full_answer.strip())
        yield "data: [DONE]\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream")

@app.get("/health")
def health():
    return {
        "status": "online",
        "hardware": get_system_stats(),
        "chunks_indexed": vector_db._collection.count()
    }

@app.get("/download-questions")
def download_questions():
    if os.path.exists(LOG_PATH):
        return FileResponse(LOG_PATH, filename="hr_questions.xlsx")
    return {"error": "No questions logged yet"}

@app.get("/download-visitors")
def download_visitors():
    if os.path.exists(VISITOR_LOG):
        return FileResponse(VISITOR_LOG, filename="visitors.xlsx")
    return {"error": "No visitors logged yet"}
